import hashlib
import random
import string
import time
import uuid
import re
from datetime import datetime, UTC
from typing import Optional, Any
from passlib.context import CryptContext

from bson import ObjectId
from fastapi import HTTPException, status

from openpyxl import load_workbook
import io

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def generate_id(prefix: str = "GEN", length: int = 8) -> str:
    """
    Generates a unique ID with a specified prefix and a UUID4-based suffix.

    Args:
        prefix (str): A string prefix for the ID (e.g., "DET" for Data Element Template).
        length (int): The desired length of the UUID part of the ID. Max 32.

    Returns:
        str: The generated unique ID.
    """
    if not 0 < length <= 32:
        raise ValueError("Length for UUID part must be between 1 and 32.")

    unique_suffix = uuid.uuid4().hex[:length].upper()
    return f"{prefix}-{unique_suffix}"


def is_valid_email(email: str) -> bool:
    """
    Checks if a string is a valid email format.

    Args:
        email (str): The email string to validate.

    Returns:
        bool: True if valid, False otherwise.
    """

    email_regex = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
    return bool(email_regex.match(email))


def get_current_utc_timestamp() -> datetime:
    """
    Returns the current UTC datetime with timezone information.

    Returns:
        datetime: Current UTC datetime.
    """
    return datetime.now(UTC)


def convert_objectid_to_str(obj: Any) -> Any:
    """
    Recursively converts MongoDB ObjectIds in a dictionary or list to strings.
    Useful before returning data from the API if Pydantic's json_encoders isn't catching everything
    or for logging purposes.

    Args:
        obj (Any): The object (dict, list, or other) to process.

    Returns:
        Any: The processed object with ObjectIds converted.
    """
    from bson import ObjectId

    if isinstance(obj, ObjectId):
        return str(obj)
    elif isinstance(obj, dict):
        return {k: convert_objectid_to_str(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_objectid_to_str(elem) for elem in obj]
    else:
        return obj


def serialize_datetime(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, list):
        return [serialize_datetime(item) for item in obj]
    if isinstance(obj, dict):
        return {key: serialize_datetime(value) for key, value in obj.items()}
    return obj


def validate_object_id(id_str: str) -> ObjectId:
    """
    Validates and returns a MongoDB ObjectId from a string.

    Args:
        id_str (str): The string to validate and convert.

    Returns:
        ObjectId: The valid ObjectId instance.

    Raises:
        ValueError: If the id_str is not a valid ObjectId.
    """

    if not ObjectId.is_valid(id_str):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid ObjectId format: {id_str}",
        )
    return ObjectId(id_str)


def hash_shake256(value: str) -> str:
    """
    Compute the SHAKE-256 hash of the provided string.
    Parameters:
        value (str): The input string to hash.
    Returns:
        str: A hexadecimal string representing the SHAKE-256 hash of the input. If the input is empty or None, returns an empty string.
    Notes:
        The function uses hashlib.shake_256, which generates a variable-length digest.
        Here, the default behavior assumes a proper hexdigest conversion.
    """

    if value:
        return hashlib.shake_256(value.encode()).hexdigest(length=32)
    else:
        return ""


def generate_blockchain_hash():

    data = "".join(random.choices(string.ascii_letters + string.digits, k=32))
    nonce = random.randint(0, 10**6)
    timestamp = str(time.time())

    block_contents = f"{data}-{nonce}-{timestamp}"

    hash_result = hashlib.sha256(block_contents.encode()).hexdigest()
    return f"0x{hash_result}"


def count_rows_in_file(contents: bytes, file_type: str) -> int:
    """
    Fast-path for CSV (newline count) and simple XLSX (openpyxl read_only).
    """
    if file_type == "csv":

        return contents.count(b"\n")
    elif file_type == "xlsx":

        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        return wb.active.max_row
    else:
        return 0


def ensure_utc(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt
